#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def criar_planilha_cronograma():
    # Criar workbook
    wb = openpyxl.Workbook()
    
    # Criar abas
    ws_cronograma = wb.active
    ws_cronograma.title = "Cronograma Detalhado"
    
    ws_checklist = wb.create_sheet("Checklist Tarefas")
    ws_apis = wb.create_sheet("APIs e Integrações")
    ws_jogos = wb.create_sheet("Jogos Detalhados")
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="9A45FC", end_color="9A45FC", fill_type="solid")
    
    subheader_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="2AD6D0", end_color="2AD6D0", fill_type="solid")
    
    priority_high = PatternFill(start_color="FFE118", end_color="FFE118", fill_type="solid")
    priority_medium = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    priority_low = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ABA 1: CRONOGRAMA DETALHADO
    ws = ws_cronograma
    
    # Cabeçalho principal
    ws.merge_cells('A1:H1')
    ws['A1'] = "🐱 CRONOGRAMA MEOW TOKEN - FOLGA DUPLA 16-17/07/2025"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="0B0019", end_color="0B0019", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers da tabela
    headers = ["Data", "Horário", "Duração", "Tarefa", "Categoria", "Prioridade", "Status", "Observações"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados do cronograma
    cronograma_data = [
        # DIA 1 - 16/07/2025
        ["16/07", "08:00-08:30", "30min", "Configuração de Secrets no Replit", "Setup", "Alta", "Pendente", "Chaves API seguras"],
        ["16/07", "08:30-09:30", "1h", "Implementação do Meow Clicker", "Jogos", "Máxima", "Pendente", "Jogo principal do projeto"],
        ["16/07", "09:30-10:30", "1h", "Sistema de energia e regeneração", "Jogos", "Alta", "Pendente", "100 pontos, regenera 1/seg"],
        ["16/07", "10:30-11:00", "30min", "Coffee break + testes", "Pausa", "Baixa", "Pendente", "Validar implementações"],
        ["16/07", "11:00-12:00", "1h", "Crypto Quiz - estrutura base", "Jogos", "Alta", "Pendente", "Sistema de perguntas"],
        ["16/07", "14:00-15:00", "1h", "Crypto Quiz - banco de perguntas", "Jogos", "Alta", "Pendente", "100+ perguntas Solana/DeFi"],
        ["16/07", "15:00-16:00", "1h", "Lucky Spin - algoritmo probabilidade", "Jogos", "Alta", "Pendente", "Jackpot progressivo"],
        ["16/07", "16:00-16:30", "30min", "Testes e ajustes", "QA", "Média", "Pendente", "Validar jogos implementados"],
        ["16/07", "16:30-17:30", "1h", "Treasure Hunt - mapas procedurais", "Jogos", "Alta", "Pendente", "Mapas 10x10"],
        ["16/07", "17:30-18:00", "30min", "Battle Arena - estrutura básica", "Jogos", "Média", "Pendente", "Combate por turnos"],
        ["16/07", "19:00-20:00", "1h", "Jupiter API - configuração", "Blockchain", "Máxima", "Pendente", "Swaps Solana"],
        ["16/07", "20:00-21:00", "1h", "Sistema de pontos blockchain", "Blockchain", "Máxima", "Pendente", "Validação on-chain"],
        ["16/07", "21:00-22:00", "1h", "Testes de integração", "QA", "Alta", "Pendente", "Validar blockchain"],
        
        # DIA 2 - 17/07/2025
        ["17/07", "08:00-09:00", "1h", "CoinGecko API - preços tempo real", "APIs", "Alta", "Pendente", "Dados de mercado"],
        ["17/07", "09:00-10:00", "1h", "Twitter API - verificação social", "APIs", "Alta", "Pendente", "Integração social"],
        ["17/07", "10:00-10:30", "30min", "Coffee break", "Pausa", "Baixa", "Pendente", "Descanso"],
        ["17/07", "10:30-11:30", "1h", "Telegram Bot - notificações", "APIs", "Média", "Pendente", "Alertas automáticos"],
        ["17/07", "11:30-12:00", "30min", "Google Analytics 4", "Analytics", "Média", "Pendente", "Métricas de uso"],
        ["17/07", "14:00-15:00", "1h", "Sistema anti-cheat", "Segurança", "Alta", "Pendente", "ML para detecção"],
        ["17/07", "15:00-16:00", "1h", "Leaderboard tempo real", "Features", "Alta", "Pendente", "Ranking dinâmico"],
        ["17/07", "16:00-17:00", "1h", "Otimizações de performance", "Otimização", "Média", "Pendente", "Velocidade < 3s"],
        ["17/07", "17:00-18:00", "1h", "Testes finais", "QA", "Alta", "Pendente", "Validação completa"],
        ["17/07", "19:00-20:00", "1h", "Deploy produção", "Deploy", "Máxima", "Pendente", "Site ao vivo"],
        ["17/07", "20:00-21:00", "1h", "Documentação final", "Docs", "Média", "Pendente", "Guias de uso"],
    ]
    
    # Preencher dados
    for row, data in enumerate(cronograma_data, 4):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Colorir por prioridade
            if col == 6:  # Coluna Prioridade
                if value == "Máxima":
                    cell.fill = priority_high
                elif value == "Alta":
                    cell.fill = priority_medium
                elif value == "Média":
                    cell.fill = priority_low
    
    # Ajustar largura das colunas
    column_widths = [12, 15, 10, 35, 15, 12, 12, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ABA 2: CHECKLIST TAREFAS
    ws = ws_checklist
    
    # Cabeçalho
    ws.merge_cells('A1:E1')
    ws['A1'] = "🎯 CHECKLIST DETALHADO - MEOW TOKEN"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="0B0019", end_color="0B0019", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    checklist_headers = ["Categoria", "Tarefa", "Subtarefas", "Status", "Prioridade"]
    for col, header in enumerate(checklist_headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados do checklist
    checklist_data = [
        ["🎮 JOGOS", "Meow Clicker", "Sistema energia, multiplicadores, animações", "❌", "Máxima"],
        ["🎮 JOGOS", "Crypto Quiz", "100+ perguntas, timer, streak, ranking", "❌", "Alta"],
        ["🎮 JOGOS", "Lucky Spin", "Algoritmo balanceado, jackpot, cooldown", "❌", "Alta"],
        ["🎮 JOGOS", "Treasure Hunt", "Mapas 10x10, dicas, raridades", "❌", "Alta"],
        ["🎮 JOGOS", "Battle Arena", "Combate turnos, matchmaking, ELO", "❌", "Média"],
        
        ["🔗 BLOCKCHAIN", "Jupiter API", "Configuração endpoints, swaps, validação", "❌", "Máxima"],
        ["🔗 BLOCKCHAIN", "Solana RPC", "Conexão mainnet, carteiras, balanços", "❌", "Máxima"],
        ["🔗 BLOCKCHAIN", "Sistema Pontos", "Validação on-chain, anti-cheat, logs", "❌", "Máxima"],
        
        ["📡 APIs", "CoinGecko", "Preços tempo real, gráficos, alertas", "❌", "Alta"],
        ["📡 APIs", "Twitter API", "Verificação follow, retweets, engagement", "❌", "Alta"],
        ["📡 APIs", "Telegram Bot", "Notificações, comandos, grupos", "❌", "Média"],
        ["📡 APIs", "Google Analytics", "Métricas uso, conversão, comportamento", "❌", "Média"],
        
        ["🛡️ SEGURANÇA", "Secrets Management", "Chaves seguras, criptografia AES-256", "❌", "Alta"],
        ["🛡️ SEGURANÇA", "Anti-cheat", "ML detecção, rate limiting, banimento", "❌", "Alta"],
        ["🛡️ SEGURANÇA", "Monitoramento", "Logs auditoria, alertas, backup", "❌", "Média"],
        
        ["🚀 DEPLOY", "Otimização", "Performance < 3s, CDN, cache", "❌", "Média"],
        ["🚀 DEPLOY", "Testes Finais", "QA completo, validação features", "❌", "Alta"],
        ["🚀 DEPLOY", "Produção", "Deploy live, domínio, SSL", "❌", "Máxima"],
    ]
    
    # Preencher checklist
    for row, data in enumerate(checklist_data, 4):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Colorir por prioridade
            if col == 5:  # Coluna Prioridade
                if value == "Máxima":
                    cell.fill = priority_high
                elif value == "Alta":
                    cell.fill = priority_medium
                elif value == "Média":
                    cell.fill = priority_low
    
    # Ajustar largura das colunas
    checklist_widths = [20, 25, 40, 10, 12]
    for i, width in enumerate(checklist_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ABA 3: APIs E INTEGRAÇÕES
    ws = ws_apis
    
    # Cabeçalho
    ws.merge_cells('A1:F1')
    ws['A1'] = "📡 APIs E INTEGRAÇÕES - CONFIGURAÇÃO DETALHADA"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="0B0019", end_color="0B0019", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    api_headers = ["API/Serviço", "Categoria", "Função", "Endpoint/Config", "Status", "Prioridade"]
    for col, header in enumerate(api_headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados das APIs
    api_data = [
        ["Jupiter API", "Blockchain", "Swaps Solana", "https://quote-api.jup.ag/v6/", "❌", "Máxima"],
        ["Phantom Wallet", "Blockchain", "Conexão carteiras", "Solana Wallet Adapter", "❌", "Máxima"],
        ["Solana RPC", "Blockchain", "Blockchain direto", "https://api.mainnet-beta.solana.com", "❌", "Máxima"],
        ["Helius RPC", "Blockchain", "Performance premium", "https://rpc.helius.xyz/", "❌", "Alta"],
        
        ["CoinGecko", "Preços", "Dados gratuitos", "https://api.coingecko.com/api/v3/", "❌", "Alta"],
        ["CoinMarketCap", "Preços", "Dados premium", "https://pro-api.coinmarketcap.com/", "❌", "Média"],
        ["Binance API", "Preços", "Tempo real", "https://api.binance.com/api/v3/", "❌", "Média"],
        
        ["Twitter API v2", "Social", "Verificação social", "https://api.twitter.com/2/", "❌", "Alta"],
        ["Telegram Bot", "Social", "Notificações", "https://api.telegram.org/bot", "❌", "Média"],
        ["Discord Webhook", "Social", "Alertas", "Discord Webhook URL", "❌", "Baixa"],
        
        ["Google reCAPTCHA", "Segurança", "Anti-bot", "https://www.google.com/recaptcha/", "❌", "Alta"],
        ["Auth0", "Segurança", "Autenticação", "https://auth0.com/", "❌", "Média"],
        
        ["Supabase", "Database", "Backend", "https://supabase.com/", "✅", "Máxima"],
        ["Google Analytics", "Analytics", "Métricas", "https://analytics.google.com/", "❌", "Média"],
        ["Mixpanel", "Analytics", "Eventos", "https://mixpanel.com/", "❌", "Baixa"],
    ]
    
    # Preencher APIs
    for row, data in enumerate(api_data, 4):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Colorir por prioridade
            if col == 6:  # Coluna Prioridade
                if value == "Máxima":
                    cell.fill = priority_high
                elif value == "Alta":
                    cell.fill = priority_medium
                elif value == "Média":
                    cell.fill = priority_low
    
    # Ajustar largura das colunas
    api_widths = [20, 15, 25, 35, 10, 12]
    for i, width in enumerate(api_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ABA 4: JOGOS DETALHADOS
    ws = ws_jogos
    
    # Cabeçalho
    ws.merge_cells('A1:E1')
    ws['A1'] = "🎮 ESPECIFICAÇÕES DETALHADAS DOS JOGOS"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="0B0019", end_color="0B0019", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    jogos_headers = ["Jogo", "Mecânica Principal", "Features", "Recompensas", "Prioridade"]
    for col, header in enumerate(jogos_headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados dos jogos
    jogos_data = [
        ["🐱 Meow Clicker", "Click para ganhar pontos", "Energia (100), regeneração (1/seg), multiplicadores (2x,5x,10x), animações, conquistas", "1-10 pontos por click", "Máxima"],
        ["🧠 Crypto Quiz", "Perguntas sobre Solana/DeFi", "100+ perguntas, timer 30s, streak system, explicações educativas, ranking", "5-50 pontos por resposta", "Alta"],
        ["🎰 Lucky Spin", "Roleta com prêmios", "Algoritmo balanceado, jackpot progressivo, cooldown 1h, efeitos visuais, histórico", "10-1000 pontos aleatórios", "Alta"],
        ["🗺️ Treasure Hunt", "Caça ao tesouro em mapas", "Mapas 10x10 procedurais, sistema dicas, raridades, inventário, temas", "20-500 pontos por tesouro", "Alta"],
        ["⚔️ Battle Arena", "Combate estratégico PvP", "Turnos, matchmaking, ranking ELO, animações, recompensas vitória", "50-200 pontos por vitória", "Média"],
    ]
    
    # Preencher jogos
    for row, data in enumerate(jogos_data, 4):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [3]:  # Colunas com texto longo
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Colorir por prioridade
            if col == 5:  # Coluna Prioridade
                if value == "Máxima":
                    cell.fill = priority_high
                elif value == "Alta":
                    cell.fill = priority_medium
                elif value == "Média":
                    cell.fill = priority_low
    
    # Ajustar largura das colunas e altura das linhas
    jogos_widths = [20, 25, 50, 25, 12]
    for i, width in enumerate(jogos_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Ajustar altura das linhas para texto longo
    for row in range(4, 9):
        ws.row_dimensions[row].height = 60
    
    # Salvar arquivo
    filename = "/home/ubuntu/cronograma_meow_token.xlsx"
    wb.save(filename)
    print(f"✅ Planilha criada com sucesso: {filename}")
    
    return filename

if __name__ == "__main__":
    criar_planilha_cronograma()

